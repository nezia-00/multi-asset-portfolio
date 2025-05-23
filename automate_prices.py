import yfinance as yf
import pandas as pd
from openpyxl import load_workbook

# === Paramètres ===
tickers = {
    'Equities': 'CW8.PA',
    'Bonds': 'AGGH',
    'Gold': 'GLD',
    'Bitcoin': 'IBIT',
    'REIT': 'REET',
    'GreenTech': 'INRG.L'
}

start_date = '2024-01-01'
end_date = pd.Timestamp.today().strftime('%Y-%m-%d')  # Aujourd'hui
excel_file = 'portfolio.xlsx'
target_sheet = 'Données'

# === 1. Télécharger les données (journalières, ajustées) ===
data = yf.download(
    tickers=list(tickers.values()),
    start=start_date,
    end=end_date,
    progress=True,
    group_by='ticker',
    auto_adjust=True
)

# === 2. Extraire prix de fin de mois ===
prices = pd.DataFrame()
for name, ticker in tickers.items():
    try:
        prices[name] = data[ticker]['Close']
    except KeyError:
        print(f"Données manquantes pour : {ticker}")

prices.index = pd.to_datetime(prices.index)
monthly_data = prices.resample('ME').last().dropna()
monthly_data.reset_index(inplace=True)

# === 3. Charger Excel existant ===
wb = load_workbook(excel_file)
ws = wb[target_sheet]

# === 4. Lire les dates existantes dans Excel ===
existing_dates = set()
for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
    if row[0]:
        existing_dates.add(pd.to_datetime(row[0]).date())

# === 5. Écrire l’en-tête si vide ===
if ws.max_row == 1:
    ws.append(["Date"] + list(tickers.keys()))

# === 6. Ajouter uniquement les nouveaux mois ===
for i in range(len(monthly_data)):
    row_date = monthly_data.loc[i, 'Date'].date()
    if row_date not in existing_dates:
        row = [row_date] + list(monthly_data.loc[i, tickers.keys()])
        ws.append(row)

# === 7. Sauvegarde ===
wb.save(excel_file)
print("Données mensuelles mises à jour dans portfolio.xlsx (feuille Données)")